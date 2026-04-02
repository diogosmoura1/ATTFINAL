import os
import re
import pandas as pd
from dicionary import empresas
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ==========================================================
# COLUNAS QUE NÃO DEVEM APARECER NO ARQUIVO FINAL
# ==========================================================
COLUNAS_REMOVER = [
    "contrato",
    "situação",
    "saldo devedor",
    "parcelas",
    "prox. venc.",
    "atraso.1",
    "dividido em",
    "faltam qnt",
    "e-mail",
    "bairro",
    "cep",
]


# ==========================================================
# FUNÇÕES BÁSICAS DE LIMPEZA E LOCALIZAÇÃO
# ==========================================================
def limpar_cpf(valor):
    """
    Remove tudo que não for número do CPF e completa com zeros à esquerda,
    deixando sempre com 11 dígitos.
    """
    if pd.isna(valor):
        return ""
    return re.sub(r"\D", "", str(valor)).zfill(11)


def encontrar_coluna_cpf(df):
    """
    Procura no DataFrame a coluna que contenha 'cpf' no nome.
    Retorna o nome real da coluna encontrada.
    """
    for col in df.columns:
        if "cpf" in str(col).strip().lower():
            return col
    return None


def encontrar_coluna_nome(df):
    """
    Procura exatamente a coluna 'nome completo'.
    Essa função é usada para inserir a coluna 'dias úteis'
    logo depois dela.
    """
    for col in df.columns:
        nome = str(col).strip().lower()
        if nome == "nome completo":
            return col
    return None


def encontrar_aba(caminho, nome_aba):
    """
    Procura uma aba no Excel ignorando diferenças de maiúsculas/minúsculas.
    Exemplo: 'histórico de atraso' e 'HISTÓRICO DE ATRASO' serão tratados como iguais.
    """
    xls = pd.ExcelFile(caminho)
    for aba in xls.sheet_names:
        if aba.strip().lower() == nome_aba.strip().lower():
            return aba
    raise Exception(f'Aba "{nome_aba}" não encontrada. Abas disponíveis: {xls.sheet_names}')


def normalizar_nome_coluna(nome):
    """
    Normaliza o nome da coluna para comparação:
    tira espaços laterais e converte para minúsculo.
    """
    return str(nome).strip().lower()


# ==========================================================
# CONVERSÃO DE VALORES MONETÁRIOS
# ==========================================================
def converter_para_float_brl(valor):
    """
    Converte textos como:
    'R$ 1.234,56' -> 1234.56

    Isso é importante para o Excel reconhecer como número,
    permitindo soma ao selecionar várias células.
    """
    if pd.isna(valor):
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = texto.replace("R$", "").replace(" ", "")

    # Exemplo: 1.234,56 -> 1234.56
    if "." in texto and "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except:
        return None


# ==========================================================
# REMOÇÃO DE COLUNAS INDESEJADAS
# ==========================================================
def remover_colunas_indesejadas(df):
    """
    Remove do DataFrame apenas as colunas listadas em COLUNAS_REMOVER,
    respeitando o nome real da coluna que veio da planilha.
    """
    mapa_colunas = {normalizar_nome_coluna(col): col for col in df.columns}
    colunas_existentes_para_remover = []

    for col in COLUNAS_REMOVER:
        col_normalizada = normalizar_nome_coluna(col)
        if col_normalizada in mapa_colunas:
            colunas_existentes_para_remover.append(mapa_colunas[col_normalizada])

    return df.drop(columns=colunas_existentes_para_remover, errors="ignore")


# ==========================================================
# TRATAMENTO DAS COLUNAS DE MOEDA
# ==========================================================
def tratar_colunas_moeda(df):
    """
    Converte as colunas monetárias para float.
    O formato visual de moeda será aplicado depois no Excel,
    mas aqui elas continuam numéricas.
    """
    mapa_colunas = {normalizar_nome_coluna(col): col for col in df.columns}

    for nome_coluna in ["valor contratado", "valor parcela"]:
        nome_normalizado = normalizar_nome_coluna(nome_coluna)

        if nome_normalizado in mapa_colunas:
            col_real = mapa_colunas[nome_normalizado]
            df[col_real] = df[col_real].apply(converter_para_float_brl)

    return df


# ==========================================================
# CRIAÇÃO DA COLUNA "DIAS ÚTEIS"
# ==========================================================
def adicionar_coluna_dias_uteis(df_evadidos, df_historico):
    """
    Conta quantas vezes o CPF aparece na aba 'histórico de atraso',
    semelhante a um CONT.SE do Excel.

    Depois adiciona a coluna 'dias úteis' logo após 'nome completo'.
    """
    col_cpf_evadidos = encontrar_coluna_cpf(df_evadidos)
    col_cpf_historico = encontrar_coluna_cpf(df_historico)
    col_nome = encontrar_coluna_nome(df_evadidos)

    if not col_cpf_evadidos:
        raise Exception("Coluna CPF não encontrada no dataframe de evadidos.")

    if not col_cpf_historico:
        raise Exception('Coluna CPF não encontrada na aba "histórico de atraso".')

    df_historico = df_historico.copy()
    df_historico[col_cpf_historico] = df_historico[col_cpf_historico].apply(limpar_cpf)
    df_historico = df_historico[df_historico[col_cpf_historico] != ""]

    # Conta quantas vezes cada CPF aparece no histórico
    contagem_cpfs = df_historico[col_cpf_historico].value_counts().to_dict()

    df_evadidos = df_evadidos.copy()
    df_evadidos["dias úteis"] = df_evadidos[col_cpf_evadidos].map(contagem_cpfs).fillna(0).astype(int)

    # Move a coluna "dias úteis" para logo depois de "nome completo"
    if col_nome:
        colunas = list(df_evadidos.columns)
        colunas.remove("dias úteis")

        indice_nome = colunas.index(col_nome)
        colunas.insert(indice_nome + 1, "dias úteis")

        df_evadidos = df_evadidos[colunas]

    return df_evadidos


# ==========================================================
# AJUSTE AUTOMÁTICO DA LARGURA DAS COLUNAS
# ==========================================================
def ajustar_largura_colunas(ws, largura_minima=10, largura_maxima=45):
    """
    Ajusta automaticamente a largura de cada coluna com base no maior
    conteúdo encontrado nela.

    largura_minima: evita colunas pequenas demais
    largura_maxima: evita colunas gigantes
    """
    for coluna in ws.columns:
        indice_coluna = coluna[0].column
        letra_coluna = get_column_letter(indice_coluna)

        maior_tamanho = 0

        for celula in coluna:
            valor = celula.value

            if valor is None:
                continue

            # Para números monetários, considera tamanho formatado aproximado
            if isinstance(valor, (int, float)):
                texto = f"{valor}"
            else:
                texto = str(valor)

            if len(texto) > maior_tamanho:
                maior_tamanho = len(texto)

        largura_ajustada = maior_tamanho + 2

        if largura_ajustada < largura_minima:
            largura_ajustada = largura_minima

        if largura_ajustada > largura_maxima:
            largura_ajustada = largura_maxima

        ws.column_dimensions[letra_coluna].width = largura_ajustada


# ==========================================================
# FORMATAÇÕES VISUAIS NO EXCEL
# ==========================================================
def aplicar_formatacao_excel(caminho_arquivo, nome_aba="Evadidos"):
    """
    Aplica:
    - formato monetário nas colunas de valor
    - cores na coluna 'dias úteis'
    - ajuste automático da largura das colunas
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]

    # Mapeia os cabeçalhos para saber em que índice está cada coluna
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col_idx).value
        if valor is not None:
            headers[str(valor).strip().lower()] = col_idx

    # Aplica formatação monetária real no Excel
    for nome_coluna in ["valor contratado", "valor parcela"]:
        if nome_coluna in headers:
            idx = headers[nome_coluna]
            for row in range(2, ws.max_row + 1):
                celula = ws.cell(row=row, column=idx)
                if celula.value not in (None, ""):
                    celula.number_format = r'"R$" #,##0.00'

    # Cores da coluna "dias úteis"
    amarelo = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    laranja = PatternFill(fill_type="solid", start_color="FFA500", end_color="FFA500")
    vermelho = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")

    if "dias úteis" in headers:
        idx_dias = headers["dias úteis"]

        for row in range(2, ws.max_row + 1):
            celula = ws.cell(row=row, column=idx_dias)
            valor = celula.value

            try:
                valor = int(valor)
            except:
                continue

            if valor > 20:
                celula.fill = vermelho
            elif valor > 10:
                celula.fill = laranja
            elif valor > 5:
                celula.fill = amarelo

    # Ajusta a largura das colunas com base no conteúdo
    ajustar_largura_colunas(ws)

    wb.save(caminho_arquivo)


# ==========================================================
# TRANSFORMA A PLANILHA EM TABELA COM FILTRO
# ==========================================================
def transformar_em_tabela_excel(caminho_arquivo, nome_aba="Evadidos"):
    """
    Transforma o intervalo de dados em uma Tabela do Excel,
    com filtro automático e estilo visual.
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]

    max_linha = ws.max_row
    max_coluna = ws.max_column

    if max_linha < 1 or max_coluna < 1:
        wb.save(caminho_arquivo)
        return

    ultima_coluna = get_column_letter(max_coluna)
    intervalo = f"A1:{ultima_coluna}{max_linha}"

    if not ws.tables:
        tabela = Table(displayName="TabelaEvadidos", ref=intervalo)

        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

    wb.save(caminho_arquivo)


# ==========================================================
# PROCESSAMENTO DE UMA EMPRESA
# ==========================================================
def processar(nome_empresa, caminho_arquivo):
    """
    Fluxo completo:
    1. Lê as abas necessárias do arquivo
    2. Remove CPFs vazios
    3. Remove duplicados da aba bcd
    4. Exclui CPFs que já estão em CARTEIRA GERAL
    5. Adiciona a coluna 'dias úteis'
    6. Converte valores monetários
    7. Remove colunas indesejadas
    8. Salva o resultado em Excel formatado
    """
    print(f"\n🔄 {nome_empresa}")

    if not os.path.exists(caminho_arquivo):
        print(f"❌ Arquivo não encontrado: {caminho_arquivo}")
        return

    try:
        # Localiza as abas necessárias dentro do arquivo da carteira
        aba_bcd = encontrar_aba(caminho_arquivo, "bcd")
        aba_carteira = encontrar_aba(caminho_arquivo, "CARTEIRA GERAL")
        aba_historico = encontrar_aba(caminho_arquivo, "histórico de atraso")

        # Lê as abas como texto para evitar conversões erradas automáticas
        df_bcd = pd.read_excel(caminho_arquivo, sheet_name=aba_bcd, dtype=str)
        df_carteira = pd.read_excel(caminho_arquivo, sheet_name=aba_carteira, dtype=str)
        df_historico = pd.read_excel(caminho_arquivo, sheet_name=aba_historico, dtype=str)

        # Descobre quais colunas são as de CPF
        col_bcd = encontrar_coluna_cpf(df_bcd)
        col_cart = encontrar_coluna_cpf(df_carteira)

        if not col_bcd:
            print("❌ Coluna CPF não encontrada na aba BCD")
            return

        if not col_cart:
            print("❌ Coluna CPF não encontrada na aba CARTEIRA GERAL")
            return

        # Limpa CPF nas duas abas principais
        df_bcd[col_bcd] = df_bcd[col_bcd].apply(limpar_cpf)
        df_carteira[col_cart] = df_carteira[col_cart].apply(limpar_cpf)

        # Remove linhas sem CPF
        df_bcd = df_bcd[df_bcd[col_bcd] != ""].copy()
        df_carteira = df_carteira[df_carteira[col_cart] != ""].copy()

        # Remove duplicados da aba bcd, mantendo apenas o primeiro CPF
        df_bcd = df_bcd.drop_duplicates(subset=col_bcd, keep="first").copy()

        # Remove da bcd todos os CPFs que já existem na carteira geral
        cpfs_carteira = set(df_carteira[col_cart])
        df_evadidos = df_bcd[~df_bcd[col_bcd].isin(cpfs_carteira)].copy()

        # Adiciona coluna "dias úteis" com base no histórico de atraso
        df_evadidos = adicionar_coluna_dias_uteis(df_evadidos, df_historico)

        # Converte colunas monetárias para número
        df_evadidos = tratar_colunas_moeda(df_evadidos)

        # Remove colunas que você não quer no resultado final
        df_evadidos = remover_colunas_indesejadas(df_evadidos)

        # Pasta de saída
        pasta_saida = os.path.normpath(
            r"C:\Users\Micro\Documents\evadidos historico"
        )
        os.makedirs(pasta_saida, exist_ok=True)

        # Nome do arquivo final
        caminho_saida = os.path.join(
            pasta_saida,
            f"EVADIDOS HISTORICOS {nome_empresa}.xlsx"
        )

        print("📁 Salvando em:", caminho_saida)

        # Salva primeiro os dados brutos tratados
        with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
            df_evadidos.to_excel(writer, index=False, sheet_name="Evadidos")

        # Aplica formatações e tabela
        aplicar_formatacao_excel(caminho_saida, nome_aba="Evadidos")
        transformar_em_tabela_excel(caminho_saida, nome_aba="Evadidos")

        print(f"✅ {len(df_evadidos)} evadidos salvos")

    except Exception as e:
        print(f"❌ Erro em {nome_empresa}: {e}")


# ==========================================================
# LOOP PRINCIPAL: PROCESSA TODAS AS EMPRESAS DO DICIONÁRIO
# ==========================================================
for nome_empresa, dados in empresas.items():
    processar(nome_empresa, dados["carteira"])